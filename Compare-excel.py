
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

def read_replace_instructions():
    """
    Đọc file text và trả về danh sách các cặp giá trị thay thế.
    """
    # Đường dẫn tới file text chứa các cặp giá trị thay thế
    replace_text_file = "replace_instructions.txt"
    replace_dict = {}
    with open(replace_text_file, 'r', encoding='utf-8') as file:
        lines = file.readlines()
        for line in lines:
            if '=' in line:
                old_value, new_value = line.strip().split('=')
                replace_dict[old_value] = new_value
    return replace_dict

def get_first_file_in_folder(folder_path):
    """
    Lấy file đầu tiên trong thư mục (theo thứ tự index).
    """
    files = os.listdir(folder_path)
    if len(files) == 0:
        raise FileNotFoundError(f"Thư mục {folder_path} không chứa file nào.")
    return os.path.join(folder_path, files[0])  # Trả về đường dẫn file đầu tiên

def unmerge_excel(file_path, header_row):
    """
    Xử lý các ô merge trong file Excel và trả về DataFrame đã xử lý.
    """
    wb = load_workbook(file_path)
    sheet = wb.worksheets[0]  # Lấy sheet đầu tiên bằng chỉ số, không cần tên

    # Lưu danh sách các vùng merge để xử lý
    merged_cells = list(sheet.merged_cells)

    for merged_cell in merged_cells:
        # Lấy tọa độ của vùng merge
        min_row, min_col, max_row, max_col = merged_cell.min_row, merged_cell.min_col, merged_cell.max_row, merged_cell.max_col
        # Lấy giá trị từ ô trên cùng bên trái của vùng merge
        top_left_value = sheet.cell(min_row, min_col).value
        # Unmerge vùng merge
        sheet.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
        # Gán giá trị vào tất cả các ô trong vùng trước đây được merge
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                sheet.cell(row, col).value = top_left_value

    # Lưu lại dữ liệu đã xử lý vào DataFrame
    data = sheet.values
    df = pd.DataFrame(data)

    # Lấy dòng header và đặt làm tên cột
    df.columns = df.iloc[header_row]  # Dòng header_row sẽ trở thành tên cột
    df = df[header_row + 1:]  # Bỏ các dòng trước header
    df.reset_index(drop=True, inplace=True)  # Reset lại index

    return df, wb, sheet  # Trả về thêm workbook và sheet để sau này tô màu trực tiếp

def compare_excel_files(excel_1_folder, excel_2_folder, result_folder, key_columns, header_row):
    """
    So sánh hai file Excel dựa trên key (các cột được chọn) và tô màu hồng các ô không khớp.
    """
    # Lấy file đầu tiên trong thư mục Excel_1 và Excel_2
    file1 = get_first_file_in_folder(excel_1_folder)
    file2 = get_first_file_in_folder(excel_2_folder)

    # Load workbook và sheet của file Excel_2
    wbResult = load_workbook(file2)
    sheetResult = wbResult.worksheets[0]

    # Xử lý các ô merge và đọc dữ liệu từ file Excel
    sheet1, _,_ = unmerge_excel(file1, header_row)
    sheet2,_,_ = unmerge_excel(file2, header_row)  

    # Bỏ các ô trống (fill NaN thành chuỗi rỗng)
    sheet1.fillna("", inplace=True)
    sheet2.fillna("", inplace=True)

    # Kiểm tra các cột key có tồn tại trong cả hai file
    for key in key_columns:
        if key not in sheet1.columns or key not in sheet2.columns:
            raise KeyError(f"Cột key '{key}' không tồn tại trong một trong hai file Excel.")

    # Tạo key để so sánh
    sheet1['key'] = sheet1[key_columns].apply(lambda row: '_'.join(row.values.astype(str)), axis=1)
    sheet2['key'] = sheet2[key_columns].apply(lambda row: '_'.join(row.values.astype(str)), axis=1)

    # Đặt key làm index để dễ so sánh
    sheet1.set_index('key', inplace=True)
    sheet2.set_index('key', inplace=True)

    # Tô màu hồng cho các ô không khớp
    pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

    # Đọc file text và lưu các cặp giá trị thay thế
    valueReplace = read_replace_instructions()

    # Đếm số lượng Cell không khớp
    countCellDiff = 0
    countRowDiff = 0
    # So sánh từng ô và tô màu
    for key in sheet2.index:
        if key in sheet1.index:  # Nếu key tồn tại trong cả hai file
            for col in sheet2.columns:
                value1 = sheet1.at[key, col] if col in sheet1.columns else ""
                value2 = sheet2.at[key, col] if col in sheet2.columns else ""
                if value2 in valueReplace:
                    value2 = valueReplace[value2]

                 #chuyển thành chuỗi để so sánh nếu NAN chuyển thành rỗng
                if pd.isnull(value1):
                    value1 = ""
                if pd.isnull(value2):
                    value2 = ""
                if str(value1) != str(value2):  # Nếu giá trị không khớp, tô màu hồng
                    row_idx = sheet2.index.get_loc(key) + header_row + 2  # Dòng trong Excel (bắt đầu từ header_row + 2 vì có header)
                    col_idx = sheet2.columns.get_loc(col) + 1  # Cột trong Excel (bắt đầu từ 1 vì không có index)
                    sheetResult.cell(row=row_idx, column=col_idx).fill = pink_fill
                    print(f"Không khớp tại dòng {row_idx}, cột {col}: {value1} vs {value2}") #In thông tin dòng và cột bị lệch
                    countCellDiff += 1
        else:
            # Nếu key chỉ tồn tại trong file Excel_2, tô màu toàn bộ dòng
            row_idx = sheet2.index.get_loc(key) + header_row + 2
            print(f"Không khớp tại dòng {row_idx}") #In thông tin dòng và cột bị lệch
            countRowDiff += 1
            for col_idx in range(1, len(sheet2.columns) + 1):  # Bắt đầu từ cột 1
                sheetResult.cell(row=row_idx, column=col_idx).fill = pink_fill
                
    # Tạo tên file kết quả với thời gian hiện tại
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    result_file = os.path.join(result_folder, f"Result_{timestamp}.xlsx")
    wbResult.save(result_file)

    #In số lượng Cell không khớp
    print(f"Số lượng Cell không khớp: {countCellDiff}")

    #In số lượng Row không khớp
    print(f"Số lượng Row không khớp: {countRowDiff}")


# Chạy chương trình
if __name__ == "__main__":
    # Đường dẫn tới thư mục chứa file Excel
    excel_1_folder = "File_Excel/Excel_1"
    excel_2_folder = "File_Excel/Excel_2"
    result_folder = "File_Excel/Result"
    replace_text_file = "replace_instructions.txt"
    # Các cột dùng làm key (bạn có thể chỉnh sửa danh sách này theo nhu cầu)
    key_columns = ["ID", "Note"]  # Ví dụ: Cột "ID" và "Note" được dùng làm key

    # Dòng chứa header (bắt đầu từ 0, ví dụ: nếu header ở dòng 2 thì header_row = 1)
    header_row = 3  # Ví dụ: Header nằm ở dòng 4 trong Excel (dòng đầu tiên là 0)

    # Gọi hàm so sánh
    compare_excel_files(excel_1_folder, excel_2_folder, result_folder, key_columns, header_row)